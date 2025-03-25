"""
This file aims to pull players's id on GGLOL based on the player list manually created.
"""

import requests
from bs4 import BeautifulSoup

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from pathlib import Path
from settings import config

DATA_DIR = Path(config("DATA_DIR"))
MANUAL_DATA_DIR = Path(config("MANUAL_DATA_DIR"))


def load_player_list():
    excel_path = MANUAL_DATA_DIR / "player_list_S14-15.xlsx"

    wb = load_workbook(excel_path, data_only=True)
    sheet = wb.active

    player_data = []
    for row in sheet.iter_rows(min_row=2):  # skip header
        cell = row[0]
        player_name = cell.value
        player_link = cell.hyperlink.target if cell.hyperlink else None
        season = row[3].value
        league = row[2].value

        if player_name and player_link:
            player_data.append((player_name, player_link, season, league))

    df = pd.DataFrame(player_data, columns=["player", "link", "season", "league"]).drop_duplicates(["player", "link"])
    df["gg_id"] = df["link"].str.split("/").str[-5]
    df["player"] = df["player"].astype(str)
    df["season"] = df["season"].astype(str)
    df["league"] = df["league"].astype(str)

    return df[["player","gg_id","season", "league"]]


def load_player_id_df():
    df = pd.read_parquet(DATA_DIR / "player_id_link_table.parquet")
    return df

# Step 3: Display matched player IDs
if __name__ == "__main__":
    
    player_df = load_player_list()

    player_df.to_parquet(DATA_DIR / "player_id_link_table.parquet")
